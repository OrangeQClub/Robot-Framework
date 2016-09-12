from openpyxl import load_workbook
class ExcelUtility(object):
    ROBOT_LIBARARY_SCOPE='Global'

    def __init__(self):
        print 'Read Cell Value in Excel file'

    def read_cell_value(self,excelfile,sheetname,columnname,rownumber):
        #Read an existing workbook
        wb = load_workbook(filename='accounts.xlsx', read_only=True)
        #Define excel sheet to read
        ws = wb['account']
        #Define Cell
        cellToRead = ''+columnname+str(rownumber)
        #read Value
        cellValue = ws[cellToRead].value
        print "Cell Value :" , cellValue
        return cellValue

